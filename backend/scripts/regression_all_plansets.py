"""V4 Regression Harness — All-Plansets mode.

Scales ``regression_v4.py`` from 3 hand-configured pilots to all 84 matched
xlsx+PDF pairs in ``2026-04-16-AI QC/_analysis/xlsx_to_planset_ties.csv``.

For each tie-table row we:

    1. Open the xlsx and auto-sniff the header row + comment/sheet/severity/
       status columns via ``app.xlsx_template_sniffer``.
    2. Synthesize a ``Pilot`` object compatible with ``regression_v4``.
    3. Reuse ``regression_v4.process_pilot`` to load ground truth and (if a
       cached V4 run exists) score against it.

Outputs land in ``backend/data/regression/all_<timestamp>/``:

    scorecard.xlsx       summary sheet + aggregate sheet + one sheet per
                          pilot that had a cached V4 run
    ground_truth.xlsx    every labeled comment from every viable xlsx
    v4_emissions.xlsx    every V4 finding from every scored pilot
    pilots_index.xlsx    all 84 pilots with truth counts, viability flags,
                          has_cached_v4 indicator
    summary.md           narrative with per-category recall rolled up
    diff.json            machine-readable aggregate diff

Usage::

    cd backend
    .venv/Scripts/python scripts/regression_all_plansets.py
    # or focus on electrical logs only:
    .venv/Scripts/python scripts/regression_all_plansets.py --electrical-only
    # or skip pilots without cached runs entirely:
    .venv/Scripts/python scripts/regression_all_plansets.py --require-cached

Gemini-free (does not call ``--mode run-v4`` ever; that's a different path).
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse harness plumbing from regression_v4. We import as a sibling script
# by manipulating sys.path — the script is designed to be run directly.
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parent))

from regression_v4 import (  # type: ignore  # noqa: E402
    Pilot,
    PilotReport,
    process_pilot,
    V4_CATEGORIES,
    CategoryScore,
    PILOTS as MANUAL_PILOTS,
)
from app.xlsx_template_sniffer import sniff_xlsx, SniffResult  # type: ignore  # noqa: E402

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BACKEND = Path(__file__).resolve().parents[1]
PROJECT = BACKEND.parent
DB_PATH = BACKEND / "data" / "planset_qc.sqlite3"
COMMENTED_DIR = PROJECT / "2026-04-16-AI QC" / "Commented PDFs"
TIES_CSV = PROJECT / "2026-04-16-AI QC" / "_analysis" / "xlsx_to_planset_ties.csv"
REGRESSION_DIR = BACKEND / "data" / "regression"

# ---------------------------------------------------------------------------
# Stage / tag heuristics
# ---------------------------------------------------------------------------

_STAGE_RULES: list[tuple[str, str]] = [
    (r"\bIFP\b",        "IFP"),
    (r"\b30\s*%",       "30"),
    (r"\b60\s*%",       "60"),
    (r"\b90\s*%",       "90"),
    (r"\bIFC\b",        "IFC"),
    (r"\bAs[- ]?Built", "AsBuilt"),
    (r"\bRedline",      "Redline"),
    (r"\bFinal\b",      "IFC"),
]


def infer_stage(xlsx_name: str, pdf_name: str) -> str:
    """Best-effort stage extraction from filename tokens."""
    hay = f"{xlsx_name} {pdf_name}"
    for pat, stage in _STAGE_RULES:
        if re.search(pat, hay, flags=re.I):
            return stage
    return "unknown"


def make_project_tag(xlsx_name: str, pdf_name: str) -> str:
    """Short identifier used for filenames / filters. Derived from shared
    tokens between the xlsx and PDF."""
    base = Path(xlsx_name).stem.lower()
    base = re.sub(r"[^a-z0-9]+", "-", base)
    base = re.sub(r"-+", "-", base).strip("-")
    return base[:40]


# ---------------------------------------------------------------------------
# Tie-table → Pilot loader
# ---------------------------------------------------------------------------


def load_ties() -> list[dict]:
    if not TIES_CSV.exists():
        raise SystemExit(f"Tie table missing: {TIES_CSV}\nRun the xlsx survey first.")
    with open(TIES_CSV, newline="") as f:
        rows = list(csv.DictReader(f))
    return rows


def is_electrical_log(xlsx_name: str) -> bool:
    """Heuristic: does the xlsx name suggest an electrical comment log?"""
    n = xlsx_name.lower()
    if "civil" in n and "electrical" not in n:
        return False
    if any(tag in n for tag in ("elec", "dnv", "ppe", "pvp", "pvsyst")):
        return True
    # Include generic review logs (owner-reviewer style) too
    if any(tag in n for tag in ("comment log", "review", "design review")):
        return True
    return False


def build_pilots(rows: list[dict], electrical_only: bool) -> list[tuple[Pilot, SniffResult, str]]:
    """Return [(Pilot, SniffResult, warning_reason_or_empty), ...].

    Skipped rows get `Pilot=None` in the result.
    """
    out: list[tuple[Pilot | None, SniffResult | None, str]] = []
    for r in rows:
        xlsx = r["xlsx_file"]
        pdf = r["best_pdf"]
        if not xlsx or not pdf:
            continue  # unmatched rows can't be scored
        if electrical_only and not is_electrical_log(xlsx):
            continue
        xlsx_path = COMMENTED_DIR / xlsx
        if not xlsx_path.exists():
            out.append((None, None, f"xlsx not found on disk: {xlsx}"))
            continue
        try:
            sniff = sniff_xlsx(xlsx_path)
        except Exception as e:
            out.append((None, None, f"sniff failed: {e}"))
            continue
        if not sniff.viable:
            out.append((None, sniff, f"not viable ({'; '.join(sniff.warnings) or 'unknown'})"))
            continue
        stage = infer_stage(xlsx, pdf)
        pilot = Pilot(
            name=Path(xlsx).stem,
            xlsx=xlsx,
            xlsx_sheet=sniff.sheet_name,
            header_row=sniff.header_row,
            sheet_col=sniff.sheet_col,           # type: ignore[arg-type]
            severity_col=sniff.severity_col,
            status_col=sniff.status_col,
            comment_col=sniff.comment_col,       # type: ignore[arg-type]
            stage=stage,
            pdf_name=pdf,
            run_id=None,
            project_tag=make_project_tag(xlsx, pdf),
            truth_mode="reviewer_comment",
            truth_caveat="",
        )
        out.append((pilot, sniff, ""))
    return out  # type: ignore[return-value]


# ---------------------------------------------------------------------------
# Rollup across all pilots
# ---------------------------------------------------------------------------


def aggregate_category_scores(reports: list[PilotReport]) -> list[CategoryScore]:
    """Sum across pilots: one CategoryScore per V4 category reflecting the
    entire corpus."""
    truth_by_cat: Counter = Counter()
    v4_total_by_cat: Counter = Counter()
    v4_nr_by_cat: Counter = Counter()

    for rep in reports:
        for t in rep.truth:
            for c in t.v4_categories:
                truth_by_cat[c] += 1
        for f in rep.v4:
            v4_total_by_cat[f.category] += 1
            if f.status in ("Fail", "Needs Review"):
                v4_nr_by_cat[f.category] += 1

    all_cats = set(truth_by_cat) | set(v4_total_by_cat) | V4_CATEGORIES
    out: list[CategoryScore] = []
    for cat in sorted(all_cats):
        t = truth_by_cat.get(cat, 0)
        vt = v4_total_by_cat.get(cat, 0)
        vnr = v4_nr_by_cat.get(cat, 0)
        has_truth = t > 0
        emitted = vnr > 0
        out.append(CategoryScore(
            category=cat,
            truth_count=t,
            v4_total=vt,
            v4_fail_or_nr=vnr,
            has_truth=has_truth,
            v4_emitted=emitted,
            matched=has_truth and emitted,
            missed=has_truth and not emitted,
            extra=(not has_truth) and emitted,
        ))
    return out


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill("solid", fgColor="DCFCE7")
MISS_FILL = PatternFill("solid", fgColor="FEE2E2")
EXTRA_FILL = PatternFill("solid", fgColor="FEF3C7")
NEUTRAL_FILL = PatternFill("solid", fgColor="F3F4F6")


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if "\n" in s:
                s = max(s.splitlines(), key=len)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = min(max_len + 2, max_width)


def write_pilots_index(out_dir: Path, pilots_info: list[tuple]) -> Path:
    """One row per tie-table entry describing viability + truth count."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PilotsIndex"
    ws.append([
        "project_tag", "xlsx", "pdf", "stage",
        "viable?", "header_row", "sheet_col", "comment_col", "severity_col", "status_col",
        "score", "warnings", "skip_reason",
    ])
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for pilot, sniff, skip in pilots_info:
        if pilot is None:
            ws.append([
                "", "", "", "",
                "N", "", "", "", "", "",
                sniff.score if sniff else "",
                "; ".join(sniff.warnings) if sniff else "",
                skip,
            ])
        else:
            ws.append([
                pilot.project_tag, pilot.xlsx, pilot.pdf_name, pilot.stage,
                "Y",
                pilot.header_row, pilot.sheet_col, pilot.comment_col,
                pilot.severity_col or "", pilot.status_col or "",
                sniff.score if sniff else "",
                "; ".join(sniff.warnings) if sniff else "",
                skip,
            ])
    _autosize(ws)
    ws.freeze_panes = "A2"
    path = out_dir / "pilots_index.xlsx"
    wb.save(path)
    return path


def write_scorecard(out_dir: Path, reports: list[PilotReport],
                    aggregate: list[CategoryScore]) -> Path:
    wb = Workbook()
    # Sheet 1: Aggregate across ALL pilots (this is the headline)
    ws = wb.active
    ws.title = "Aggregate"
    ws.append(["Category", "Truth (all pilots)", "V4 total", "V4 Fail+NR",
               "Match?", "Miss?", "Extra?", "Verdict"])
    for c in range(1, 9):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for s in aggregate:
        verdict = []
        if s.matched: verdict.append("MATCH")
        if s.missed: verdict.append("MISS")
        if s.extra: verdict.append("EXTRA")
        if not verdict: verdict.append("-")
        ws.append([s.category, s.truth_count, s.v4_total, s.v4_fail_or_nr,
                   "Y" if s.matched else "",
                   "Y" if s.missed else "",
                   "Y" if s.extra else "",
                   ",".join(verdict)])
        last = ws.max_row
        fill = NEUTRAL_FILL
        if s.matched: fill = MATCH_FILL
        elif s.missed: fill = MISS_FILL
        elif s.extra: fill = EXTRA_FILL
        for c in range(1, 9):
            ws.cell(row=last, column=c).fill = fill
    _autosize(ws)
    ws.freeze_panes = "A2"

    # Sheet 2: Per-pilot summary
    summary = wb.create_sheet("PerPilot")
    summary.append([
        "Pilot", "project_tag", "Run ID", "Stage",
        "Truth", "V4 total", "V4 NR+Fail",
        "Cats matched", "Cats missed", "Cats extra", "Recall %", "Notes",
    ])
    for c in range(1, summary.max_column + 1):
        cell = summary.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        truth_cats = sum(1 for s in rep.scores if s.has_truth)
        matched = sum(1 for s in rep.scores if s.matched)
        missed = sum(1 for s in rep.scores if s.missed)
        extra = sum(1 for s in rep.scores if s.extra)
        recall = (matched / truth_cats * 100) if truth_cats else 0.0
        v4_total = sum(s.v4_total for s in rep.scores)
        v4_nr = sum(s.v4_fail_or_nr for s in rep.scores)
        summary.append([
            rep.pilot.name, rep.pilot.project_tag,
            (rep.run_id_used or "")[:12], rep.pilot.stage,
            len(rep.truth), v4_total, v4_nr,
            matched, missed, extra, round(recall, 1),
            "; ".join(rep.notes),
        ])
    _autosize(summary)
    summary.freeze_panes = "A2"

    # Per-pilot detail sheets — only for pilots with cached V4 runs
    scored = [r for r in reports if r.run_id_used]
    # cap to 40 sheets to keep the workbook openable
    for rep in scored[:40]:
        safe = re.sub(r"[^A-Za-z0-9 ]+", "", rep.pilot.project_tag)[:28]
        ws = wb.create_sheet(title=safe or "pilot")
        ws.append(["Category", "Truth", "V4 total", "V4 Fail+NR", "Verdict"])
        for c in range(1, 6):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for s in rep.scores:
            verdict = []
            if s.matched: verdict.append("MATCH")
            if s.missed: verdict.append("MISS")
            if s.extra: verdict.append("EXTRA")
            if not verdict: verdict.append("-")
            ws.append([s.category, s.truth_count, s.v4_total,
                       s.v4_fail_or_nr, ",".join(verdict)])
        _autosize(ws)
        ws.freeze_panes = "A2"

    path = out_dir / "scorecard.xlsx"
    wb.save(path)
    return path


def write_ground_truth(out_dir: Path, reports: list[PilotReport]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "GroundTruth"
    ws.append(["project_tag", "pilot", "xlsx", "xlsx_row", "comment_id",
               "sheet_raw", "v4_categories", "severity", "status", "stage", "text"])
    for c in range(1, 12):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        for t in rep.truth:
            ws.append([
                rep.pilot.project_tag, t.pilot, t.xlsx, t.row, t.comment_id,
                t.sheet_raw, ", ".join(t.v4_categories) or "(unmapped)",
                t.severity, t.status, t.stage, t.text,
            ])
            row = ws.max_row
            if not t.v4_categories:
                for c in range(1, 12):
                    ws.cell(row=row, column=c).fill = EXTRA_FILL
    _autosize(ws, max_width=80)
    ws.column_dimensions["K"].width = 100
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=11).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    path = out_dir / "ground_truth.xlsx"
    wb.save(path)
    return path


def write_v4_emissions(out_dir: Path, reports: list[PilotReport]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "V4_Emissions"
    ws.append(["project_tag", "pilot", "run_id", "category", "rule_key",
               "severity", "status", "page", "title", "description"])
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        for f in rep.v4:
            ws.append([
                rep.pilot.project_tag, f.pilot, (f.run_id or "")[:12],
                f.category, f.rule_key, f.severity, f.status,
                f.page_number, f.title, f.description,
            ])
    _autosize(ws, max_width=80)
    ws.freeze_panes = "A2"
    path = out_dir / "v4_emissions.xlsx"
    wb.save(path)
    return path


def write_summary_md(out_dir: Path, reports: list[PilotReport],
                     aggregate: list[CategoryScore],
                     pilots_info: list[tuple], electrical_only: bool) -> Path:
    n_total = len(pilots_info)
    n_viable = sum(1 for p, _, _ in pilots_info if p is not None)
    n_scored = sum(1 for r in reports if r.run_id_used)
    total_truth = sum(len(r.truth) for r in reports)
    total_v4 = sum(len(r.v4) for r in reports)
    total_v4_nr = sum(1 for r in reports for f in r.v4
                       if f.status in ("Fail", "Needs Review"))

    truth_cats = sum(1 for s in aggregate if s.has_truth)
    matched = sum(1 for s in aggregate if s.matched)
    missed = sum(1 for s in aggregate if s.missed)
    extra = sum(1 for s in aggregate if s.extra)
    recall = (matched / truth_cats * 100) if truth_cats else 0.0

    lines = [
        "# V4 Regression — All Plansets",
        "",
        f"_Run timestamp_: `{datetime.now(timezone.utc).isoformat()}`",
        f"_Scope_: {'electrical-only' if electrical_only else 'all matched logs'}",
        f"_Tie table_: `2026-04-16-AI QC/_analysis/xlsx_to_planset_ties.csv`",
        "",
        "## Corpus overview",
        "",
        f"- Tie-table candidates considered: **{n_total}**",
        f"- Viable pilots (auto-sniffer found comment + sheet columns): **{n_viable}**",
        f"- Scored pilots (viable AND cached V4 run exists): **{n_scored}**",
        f"- Reviewer comments ingested: **{total_truth}**",
        f"- V4 findings loaded (scored pilots): **{total_v4}** (NR+Fail: {total_v4_nr})",
        "",
        "## Aggregate headline",
        "",
        f"Across scored pilots, V4 hit **{matched}/{truth_cats}** categories the reviewers flagged.",
        "",
        f"- **Matched categories**: {matched}",
        f"- **Missed categories** (reviewers flagged, V4 silent): {missed}",
        f"- **Extra categories** (V4 flagged, reviewers silent): {extra}",
        f"- **Category-level recall**: **{recall:.1f}%**",
        "",
        "## Aggregate per-category breakdown",
        "",
        "| Category | Truth | V4 total | V4 NR+Fail | Verdict |",
        "|---|---|---|---|---|",
    ]
    for s in aggregate:
        bits = []
        if s.matched: bits.append("MATCH")
        if s.missed: bits.append("MISS")
        if s.extra: bits.append("EXTRA")
        if not bits: continue
        lines.append(f"| {s.category} | {s.truth_count} | {s.v4_total} | "
                     f"{s.v4_fail_or_nr} | {', '.join(bits)} |")
    lines.append("")

    # Top-10 most-truth-heavy categories (drives the demo narrative)
    ranked = sorted((s for s in aggregate if s.truth_count > 0),
                    key=lambda s: s.truth_count, reverse=True)[:10]
    lines.append("## Top categories by reviewer volume")
    lines.append("")
    lines.append("| Rank | Category | Truth comments | V4 hits | V4 silence? |")
    lines.append("|---|---|---|---|---|")
    for i, s in enumerate(ranked, 1):
        silence = "YES — investigate" if s.missed else "no"
        lines.append(f"| {i} | {s.category} | {s.truth_count} | {s.v4_fail_or_nr} | {silence} |")
    lines.append("")

    # Pilots awaiting analysis — these are the interesting next V4 runs
    unscored = [r for r in reports if not r.run_id_used and r.truth]
    if unscored:
        lines.append("## Pilots with ground truth but no cached V4 run")
        lines.append("")
        lines.append(f"{len(unscored)} pilots have labeled reviewer comments but no "
                     "V4 run on disk. Running V4 on these would close the biggest "
                     "coverage gap:")
        lines.append("")
        lines.append("| Pilot | Stage | Truth comments |")
        lines.append("|---|---|---|")
        for rep in sorted(unscored, key=lambda r: len(r.truth), reverse=True)[:20]:
            lines.append(f"| {rep.pilot.name} | {rep.pilot.stage} | {len(rep.truth)} |")
        lines.append("")

    lines.append("## How to read this")
    lines.append("")
    lines.append(
        "- **Scored pilots** are those with *both* a viable xlsx and a cached V4 run "
        "in SQLite. For the rest, the ground truth is ingested but there's nothing "
        "to score it against yet.")
    lines.append(
        "- **Aggregate** rows collapse every scored pilot's per-category data into "
        "one row per V4 category. MISS across the aggregate is a stronger signal "
        "than any single-pilot MISS — it means V4 is consistently silent on a "
        "category reviewers care about.")
    lines.append(
        "- The `pilots_index.xlsx` workbook lists every tie-table candidate, "
        "its viability, and why skipped rows were dropped. Use it to spot-check "
        "the auto-sniffer.")

    path = out_dir / "summary.md"
    path.write_text("\n".join(lines))
    return path


def write_diff_json(out_dir: Path, reports: list[PilotReport],
                    aggregate: list[CategoryScore]) -> Path:
    data = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "pilot_count": len(reports),
        "scored_pilot_count": sum(1 for r in reports if r.run_id_used),
        "aggregate": [asdict(s) for s in aggregate],
        "pilots": [
            {
                "name": r.pilot.name,
                "project_tag": r.pilot.project_tag,
                "stage": r.pilot.stage,
                "run_id": r.run_id_used,
                "truth_count": len(r.truth),
                "v4_count": len(r.v4),
                "notes": r.notes,
                "categories": [asdict(s) for s in r.scores if s.truth_count or s.v4_fail_or_nr],
            }
            for r in reports
        ],
    }
    path = out_dir / "diff.json"
    path.write_text(json.dumps(data, indent=2))
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--electrical-only", action="store_true",
                    help="Skip civil-only and non-electrical review logs.")
    ap.add_argument("--require-cached", action="store_true",
                    help="Include only pilots with a cached V4 run in SQLite.")
    ap.add_argument("--out-dir", default=None,
                    help="Override output directory.")
    ap.add_argument("--max-pilots", type=int, default=None,
                    help="Cap for smoke-testing; omit for full run.")
    args = ap.parse_args()

    ties = load_ties()
    pilots_info = build_pilots(ties, electrical_only=args.electrical_only)
    print(f"[*] Tie-table rows: {len(ties)}; candidates after filter: {len(pilots_info)}")
    viable = [(p, s, skip) for (p, s, skip) in pilots_info if p is not None]
    print(f"[*] Viable pilots (comment+sheet cols detected): {len(viable)}")

    # Merge in hand-configured pilots (Bishop, Cottonwood, Sawyer) from
    # regression_v4.PILOTS. These are NOT in the tie-table CSV but have
    # carefully-tuned truth loaders and (in Bishop's case) a cached V4 run,
    # so they belong in the aggregate headline.
    tag_seen = {p.project_tag for p, _, _ in viable}
    merged_manual = 0
    for mp in MANUAL_PILOTS:
        if mp.project_tag in tag_seen:
            continue
        viable.append((mp, None, "manual (regression_v4.PILOTS)"))
        pilots_info.append((mp, None, "manual (regression_v4.PILOTS)"))
        merged_manual += 1
    if merged_manual:
        print(f"[*] Merged {merged_manual} manually-configured pilots "
              f"(Bishop/Cottonwood/Sawyer).")

    if args.max_pilots:
        viable = viable[: args.max_pilots]
        print(f"[*] Capped to {len(viable)} pilots for smoke test.")

    if args.require_cached and DB_PATH.exists():
        # Prune to those with cached runs first
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        filtered: list[tuple] = []
        for p, s, skip in viable:
            cur = conn.execute(
                "SELECT id FROM runs WHERE original_filename = ? "
                "OR original_filename LIKE ? ORDER BY created_at DESC LIMIT 1",
                (p.pdf_name, f"%{p.pdf_name}%"),
            )
            if cur.fetchone():
                filtered.append((p, s, skip))
        conn.close()
        print(f"[*] After require-cached: {len(filtered)} pilots.")
        viable = filtered

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir) if args.out_dir else REGRESSION_DIR / f"all_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    reports: list[PilotReport] = []
    for pilot, sniff, _skip in viable:
        rep = process_pilot(pilot, mode="use-cached")
        reports.append(rep)
        flag = "*" if rep.run_id_used else " "
        print(f"  [{flag}] {pilot.project_tag[:40]:40s} truth={len(rep.truth):3d} v4={len(rep.v4):3d}")

    aggregate = aggregate_category_scores(reports)

    scorecard = write_scorecard(out_dir, reports, aggregate)
    gt = write_ground_truth(out_dir, reports)
    em = write_v4_emissions(out_dir, reports)
    idx = write_pilots_index(out_dir, pilots_info)
    sm = write_summary_md(out_dir, reports, aggregate, pilots_info,
                          electrical_only=args.electrical_only)
    dj = write_diff_json(out_dir, reports, aggregate)

    print()
    print(f"[OK] Output: {out_dir}")
    print(f"     scorecard     : {scorecard.name}")
    print(f"     ground truth  : {gt.name}")
    print(f"     v4 emissions  : {em.name}")
    print(f"     pilots index  : {idx.name}")
    print(f"     summary.md    : {sm.name}")
    print(f"     diff.json     : {dj.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
_scorecard(out_dir, reports, aggregate)
    gt = write_ground_truth(out_dir, reports)
    em = write_v4_emissions(out_dir, reports)
    idx = write_pilots_index(out_dir, pilots_info)
    sm = write_summary_md(out_dir, reports, aggregate, pilots_info,
                          electrical_only=args.electrical_only)
    dj = write_diff_json(out_dir, reports, aggregate)

    print()
    print(f"[OK] Output: {out_dir}")
    print(f"     scorecard     : {scorecard.name}")
    print(f"     ground truth  : {gt.name}")
    print(f"     v4 emissions  : {em.name}")
    print(f"     pilots index  : {idx.name}")
    print(f"     summary.md    : {sm.name}")
    print(f"     diff.json     : {dj.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
