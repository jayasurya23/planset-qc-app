"""V4 Regression Harness
========================

Compare V4 engine emissions against reviewer/customer comments captured in the
xlsx comment logs at ``2026-04-16-AI QC/Commented PDFs/``.

Ground-truth source: ~3,139 labeled reviewer comments across 104 xlsx logs.
See ``2026-04-16-AI QC/_analysis/README.md`` for the full survey.

Goal: for each pilot PDF, produce a precision/recall table per V4 category
that shows whether the V4 engine flagged at least one issue where a
reviewer did, and where it emitted extras or missed real findings.

Usage
-----
    cd backend
    # Dry-run: load ground truth + show category coverage, no V4 execution.
    .venv/Scripts/python scripts/regression_v4.py --mode truth-only

    # Use already-analyzed runs from SQLite.
    .venv/Scripts/python scripts/regression_v4.py --mode use-cached

    # Run V4 (Gemini calls) on any pilot missing a cached run.
    RULES_FILE=rules_v4_draft.yaml \\
        .venv/Scripts/python scripts/regression_v4.py --mode run-v4

Outputs land in ``backend/data/regression/<timestamp>/``:
    scorecard.xlsx       per-pilot, per-category precision/recall
    ground_truth.xlsx    normalized reviewer comments
    v4_emissions.xlsx    V4 findings we scored against
    diff.json            matched / missed / extra for every (pilot, category)
    summary.md           human-readable narrative for the demo

The pilots live in ``PILOTS`` below — edit that list to add/remove cases.
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BACKEND = Path(__file__).resolve().parents[1]
PROJECT = BACKEND.parent
DB_PATH = BACKEND / "data" / "planset_qc.sqlite3"
COMMENTED_DIR = PROJECT / "2026-04-16-AI QC" / "Commented PDFs"
REGRESSION_DIR = BACKEND / "data" / "regression"

# ---------------------------------------------------------------------------
# V4 taxonomy — kept in sync with rules_v4_draft.yaml
# ---------------------------------------------------------------------------

V4_CATEGORIES: set[str] = {
    "AI Input Gate", "BOD / Due Diligence", "Cross-Sheet", "Title Block",
    "E-001", "E-002", "E-010", "E-011", "E-050",
    "E-100", "E-101/E-102", "E-103/E-104", "E-106", "E-107", "E-110",
    "E-120", "E-130", "E-140",
    "E-200", "E-210", "E-214-E-217",
    "E-300", "E-400", "E-420-E-422", "E-450",
    "E-500-E-504", "E-601", "E-900",
}

# Severity normalization. Input comes from reviewer-written columns
# (Severity / Priority / col values) which vary across templates.
SEVERITY_ORDER = {"high": 3, "medium": 2, "low": 1, "note": 0}

# ---------------------------------------------------------------------------
# Pilot configuration
# ---------------------------------------------------------------------------


@dataclass
class Pilot:
    """One (xlsx log, planset PDF, optional cached V4 run) tuple."""
    name: str               # human-readable label for reports
    xlsx: str               # filename under Commented PDFs/
    xlsx_sheet: str | None  # which sheet to read; None = first
    header_row: int         # 1-indexed header row in that sheet
    sheet_col: str          # header text that identifies the Sheet Page column
    severity_col: str | None
    status_col: str | None
    comment_col: str        # the reviewer-comment column (PPE / PVP / Owner...)
    stage: str              # "30" / "60" / "90" / "IFC" / "AsBuilt"
    pdf_name: str           # original_filename in SQLite runs table
    run_id: str | None      # specific cached run to score against; None = pick newest
    project_tag: str        # used in filenames
    truth_mode: str = "reviewer_comment"  # reviewer_comment | designer_response | mixed
    truth_caveat: str = ""  # surfaced in summary.md under pilot header


PILOTS: list[Pilot] = [
    # Pilot 1 — cached V4 run exists; smallest risk of harness blow-up.
    # NOTE: Bishop xlsx is a DESIGNER RESPONSE LOG, not a reviewer comment log.
    # Each row = one sheet Castillo addressed. Recall is still meaningful at
    # sheet/category granularity (V4 should flag something where the reviewer
    # did), but per-comment text similarity is not possible without the
    # matching redline PDF ("PLEASE REFER TO REDLINE DRAWING" — per xlsx row 1).
    Pilot(
        name="Bishop 90% Electrical",
        xlsx="Bishop ELE Revision log_9.30.34 (1).xlsx",
        xlsx_sheet="Sheet1",
        header_row=4,
        sheet_col="PAGE NUMBER",
        severity_col=None,
        status_col=None,
        comment_col="RESPONDS TO COMMENTSS",  # (sic) typo in source file
        stage="90",
        pdf_name="bishop_90.pdf",
        run_id=None,  # newest wins
        project_tag="bishop",
        truth_mode="designer_response",
        truth_caveat=(
            "Bishop xlsx is a designer response log — each row is a sheet "
            "Castillo confirmed a fix on, not an original reviewer comment. "
            "Category-level recall is valid (the sheet was flagged by the "
            "reviewer), but per-comment text matching requires the redline PDF."
        ),
    ),
    # Pilot 2 — xlsx is rich, need to analyze PDF or skip if --mode truth-only.
    Pilot(
        name="Cottonwood IFC Electrical",
        xlsx="Cottonwood-IFC-Electrical-Sealed-6-05-2025-PPE and DNV Review Comments.xlsx",
        xlsx_sheet="Sheet1",
        header_row=6,
        sheet_col="Sheet Page / Location on Document",
        severity_col="Severity",
        status_col="Status",
        comment_col="PPE Engineer Comment 1",
        stage="IFC",
        pdf_name="Cottonwood.pdf",
        run_id=None,
        project_tag="cottonwood",
    ),
    # Pilot 3 — Sawyer electrical, deepest labeled dataset.
    Pilot(
        name="Sawyer IFP Electrical",
        xlsx="PPE-09276.01 - Generate - Sawyer-Electrical.xlsx",
        xlsx_sheet=None,  # first sheet
        header_row=1,
        sheet_col="Sheet Page / Location on Document",
        severity_col="Severity",
        status_col="Status",
        comment_col="PPE Engineer Comment 1",
        stage="60",
        pdf_name="PPE-09276.01 - Generate - Sawyer-Elec.pdf",
        run_id=None,
        project_tag="sawyer",
    ),
]

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------


@dataclass
class GroundTruthComment:
    pilot: str
    xlsx: str
    row: int
    comment_id: str
    sheet_raw: str
    v4_categories: list[str]   # one sheet may fan out to multiple V4 cats
    severity: str              # high/medium/low/note
    status: str                # open/closed/unknown
    stage: str
    text: str


@dataclass
class V4Finding:
    pilot: str
    run_id: str
    issue_id: str
    category: str
    rule_key: str
    severity: str
    status: str                # Pass/Fail/Needs Review/Deferred/Overridden
    title: str
    description: str
    page_number: int | None


@dataclass
class CategoryScore:
    category: str
    truth_count: int
    v4_total: int
    v4_fail_or_nr: int         # Fail + Needs Review (the "emitted an issue" signal)
    has_truth: bool
    v4_emitted: bool
    matched: bool              # truth exists AND v4 emitted
    missed: bool               # truth exists AND no v4 emission
    extra: bool                # no truth but v4 emitted


@dataclass
class PilotReport:
    pilot: Pilot
    truth: list[GroundTruthComment] = field(default_factory=list)
    v4: list[V4Finding] = field(default_factory=list)
    scores: list[CategoryScore] = field(default_factory=list)
    run_id_used: str | None = None
    notes: list[str] = field(default_factory=list)

    def truth_category_summary(self) -> dict[str, int]:
        out: dict[str, int] = {}
        for t in self.truth:
            for c in t.v4_categories:
                out[c] = out.get(c, 0) + 1
        return out


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

_SHEET_CLEAN_RE = re.compile(r"[^A-Za-z0-9./-]+")


def _clean_sheet(raw: str) -> str:
    return _SHEET_CLEAN_RE.sub("", (raw or "")).upper().strip()


def _nearest_hundred_v4_category(n: int) -> str | None:
    """Given an E-sheet number outside direct V4 coverage (e.g. 111, 302),
    return the V4 category for its nearest containing hundred, or None if
    that hundred isn't a V4 concept either.

    Mapping rules:
      110-120       -> E-110  (grounding / metering family)
      200-213       -> E-200
      214-217       -> E-214-E-217  (covered by direct-range above already)
      300-399       -> E-300
      400-419       -> E-400
      420-422       -> E-420-E-422 (covered above)
      423-499       -> E-400
      505-599       -> E-500-E-504  (trenching / civil-electrical details)
      600-699       -> E-601
      other         -> None (no rollup target in V4)
    """
    if 100 <= n <= 109:
        return "E-100"          # E-105 etc. -> general electrical layout family
    if 110 <= n <= 120:
        return "E-110"
    if 130 <= n <= 139:
        return "E-130"
    if 140 <= n <= 149:
        return "E-140"
    if 200 <= n <= 209:
        return "E-200"
    if 210 <= n <= 213:
        return "E-210"
    if 300 <= n <= 399:
        return "E-300"
    if 400 <= n <= 419 or 423 <= n <= 449:
        return "E-400"
    if 450 <= n <= 499:
        return "E-450"
    if 505 <= n <= 599:
        return "E-500-E-504"
    if 600 <= n <= 799:
        return "E-601"          # utility interconnect / pole & overhead details
    if 900 <= n <= 999:
        return "E-900"
    return None


def map_sheet_to_v4_category(sheet_raw: str) -> list[str]:
    """Resolve a reviewer-entered sheet page (e.g. 'E-101', 'E-500 & E-501',
    'G-100', 'E-100 & E-103') to one or more V4 category keys.

    Returns an empty list if we can't map it (caller treats as 'unmapped').
    """
    if not sheet_raw:
        return []
    s = str(sheet_raw).strip()

    # Common delimiters reviewers use: '&', ',', 'and', '/'
    parts = re.split(r"\s*(?:&|,|/|\band\b|\bto\b|-through-)\s*", s, flags=re.I)
    parts = [p.strip() for p in parts if p.strip()]

    # Single token? still run through the mapper as a list of one.
    if not parts:
        parts = [s]

    out: list[str] = []
    for p in parts:
        cleaned = _clean_sheet(p)
        if not cleaned:
            continue

        # Direct V4 category hit?
        if cleaned in V4_CATEGORIES:
            out.append(cleaned)
            continue

        # Normalize letter-number: E101 -> E-101, G100 -> G-100
        m = re.fullmatch(r"([A-Z]+)-?(\d{1,4})([A-Z]?(?:\.\d+)?)", cleaned)
        if not m:
            continue
        prefix, num, suffix = m.group(1), m.group(2), m.group(3)

        # G-sheets (general / title / cover / index) roll up to Title Block.
        if prefix == "G":
            out.append("Title Block")
            continue

        # Non-E sheets (C-xxx civil, M-xxx mechanical, etc.) aren't in V4.
        if prefix != "E":
            continue

        n = int(num)

        # Apply V4's slash/range groupings.
        if n in (101, 102):
            out.append("E-101/E-102")
        elif n in (103, 104):
            out.append("E-103/E-104")
        elif 214 <= n <= 217:
            out.append("E-214-E-217")
        elif 420 <= n <= 422:
            out.append("E-420-E-422")
        elif 500 <= n <= 504:
            out.append("E-500-E-504")
        else:
            code = f"E-{n:03d}"
            if code in V4_CATEGORIES:
                out.append(code)
            else:
                # Fallback: roll up to the nearest E-hundred if that hundred is
                # a V4 category (or governs a V4 range). Captures things like
                # E-111/E-112 -> E-110, E-202 -> E-200, E-301/302 -> E-300.
                # Comments attributed to specific detail sheets will thus score
                # against the parent category that houses V4's rule.
                parent = _nearest_hundred_v4_category(n)
                if parent:
                    out.append(parent)

    # de-dup while preserving first-seen order
    seen: set[str] = set()
    result: list[str] = []
    for c in out:
        if c not in seen:
            seen.add(c)
            result.append(c)
    return result


def map_severity(raw) -> str:
    if raw is None:
        return "low"
    s = str(raw).strip().lower()
    if not s:
        return "low"
    if s in ("high", "hi", "h", "critical", "crit", "major", "blocker"):
        return "high"
    if s in ("medium", "med", "m", "moderate"):
        return "medium"
    if s in ("note", "notes", "info", "informational", "notice", "comment"):
        return "note"
    return "low"


def normalize_status(raw) -> str:
    if raw is None:
        return "unknown"
    s = str(raw).strip().lower()
    if not s:
        return "unknown"
    if "close" in s or "resolved" in s or "accept" in s:
        return "closed"
    if "open" in s or "pending" in s or "wip" in s:
        return "open"
    return "unknown"


# ---------------------------------------------------------------------------
# xlsx ingestion — ground truth
# ---------------------------------------------------------------------------


def _read_header_row(ws, header_row: int) -> list[str]:
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=header_row, max_col=40), start=1):
        if i == header_row:
            headers = [str(v).strip() if v is not None else f"col{c}"
                       for c, v in enumerate(row, 1)]
            break
    return headers


def _col_index_match(headers: list[str], target: str, fuzzy: bool = True) -> int | None:
    """Find the 0-based column index whose header matches `target`."""
    if target is None:
        return None
    t = target.strip().lower()
    for i, h in enumerate(headers):
        if h.strip().lower() == t:
            return i
    if not fuzzy:
        return None
    # substring fallback
    for i, h in enumerate(headers):
        if t in h.strip().lower():
            return i
    return None


def load_ground_truth(pilot: Pilot) -> list[GroundTruthComment]:
    xlsx_path = COMMENTED_DIR / pilot.xlsx
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Missing xlsx for pilot {pilot.name}: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sn = pilot.xlsx_sheet or wb.sheetnames[0]
    if sn not in wb.sheetnames:
        raise ValueError(f"Sheet '{sn}' not in {pilot.xlsx}; found {wb.sheetnames}")
    ws = wb[sn]

    headers = _read_header_row(ws, pilot.header_row)

    sheet_idx = _col_index_match(headers, pilot.sheet_col)
    comment_idx = _col_index_match(headers, pilot.comment_col)
    severity_idx = _col_index_match(headers, pilot.severity_col) if pilot.severity_col else None
    status_idx = _col_index_match(headers, pilot.status_col) if pilot.status_col else None
    id_idx = _col_index_match(headers, "Comment ID") or _col_index_match(headers, "Comment Number")

    if sheet_idx is None or comment_idx is None:
        wb.close()
        raise ValueError(
            f"Could not find required columns for pilot {pilot.name}. "
            f"Headers: {headers}\nWanted sheet_col={pilot.sheet_col!r}, "
            f"comment_col={pilot.comment_col!r}"
        )

    results: list[GroundTruthComment] = []
    for i, row in enumerate(ws.iter_rows(values_only=True, max_col=40), start=1):
        if i <= pilot.header_row:
            continue
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        sheet_raw = row[sheet_idx] if sheet_idx < len(row) else None
        comment = row[comment_idx] if comment_idx < len(row) else None
        if comment is None or not str(comment).strip():
            continue
        v4_cats = map_sheet_to_v4_category(str(sheet_raw) if sheet_raw else "")
        sev = map_severity(row[severity_idx] if severity_idx is not None and severity_idx < len(row) else None)
        status = normalize_status(row[status_idx] if status_idx is not None and status_idx < len(row) else None)
        cid = str(row[id_idx]) if id_idx is not None and id_idx < len(row) and row[id_idx] is not None else ""

        results.append(GroundTruthComment(
            pilot=pilot.name,
            xlsx=pilot.xlsx,
            row=i,
            comment_id=cid,
            sheet_raw=str(sheet_raw) if sheet_raw is not None else "",
            v4_categories=v4_cats,
            severity=sev,
            status=status,
            stage=pilot.stage,
            text=str(comment)[:1200].strip(),
        ))
    wb.close()
    return results


# ---------------------------------------------------------------------------
# V4 emissions — load from SQLite
# ---------------------------------------------------------------------------


def find_run_id(conn: sqlite3.Connection, pilot: Pilot) -> str | None:
    """Return the run_id to score against, or None if no cached run exists."""
    if pilot.run_id:
        return pilot.run_id
    cur = conn.execute(
        "SELECT id FROM runs WHERE original_filename = ? ORDER BY created_at DESC LIMIT 1",
        (pilot.pdf_name,),
    )
    row = cur.fetchone()
    if row:
        return row["id"]
    # Fallback: LIKE match.
    cur = conn.execute(
        "SELECT id FROM runs WHERE original_filename LIKE ? ORDER BY created_at DESC LIMIT 1",
        (f"%{pilot.pdf_name}%",),
    )
    row = cur.fetchone()
    return row["id"] if row else None


def load_v4_findings(pilot: Pilot, run_id: str) -> list[V4Finding]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM issues WHERE run_id = ? ORDER BY category, page_number",
        (run_id,),
    ).fetchall()
    conn.close()
    return [
        V4Finding(
            pilot=pilot.name,
            run_id=run_id,
            issue_id=r["id"],
            category=r["category"] or "",
            rule_key=r["item_key"] or "",
            severity=(r["severity"] or "").lower(),
            status=r["status"] or "",
            title=r["title"] or "",
            description=(r["description"] or "")[:500],
            page_number=r["page_number"],
        )
        for r in rows
    ]


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------


def score(report: PilotReport) -> None:
    """Populate report.scores per category."""
    # Truth count per V4 category (a comment on E-101 contributes to E-101/E-102)
    truth_counts: dict[str, int] = {}
    for t in report.truth:
        for c in t.v4_categories:
            truth_counts[c] = truth_counts.get(c, 0) + 1

    # V4 counts per category
    v4_total: dict[str, int] = {}
    v4_fail_or_nr: dict[str, int] = {}
    for f in report.v4:
        v4_total[f.category] = v4_total.get(f.category, 0) + 1
        if f.status in ("Fail", "Needs Review"):
            v4_fail_or_nr[f.category] = v4_fail_or_nr.get(f.category, 0) + 1

    # Union of categories we saw on either side
    all_cats = set(truth_counts) | set(v4_total) | V4_CATEGORIES

    scores: list[CategoryScore] = []
    for cat in sorted(all_cats):
        t = truth_counts.get(cat, 0)
        vt = v4_total.get(cat, 0)
        vnr = v4_fail_or_nr.get(cat, 0)
        has_truth = t > 0
        emitted = vnr > 0
        scores.append(CategoryScore(
            category=cat,
            truth_count=t,
            v4_total=vt,
            v4_fail_or_nr=vnr,
            has_truth=has_truth,
            v4_emitted=emitted,
            matched=has_truth and emitted,
            missed=has_truth and not emitted,
            extra=(not has_truth) and emitted,
        ))
    report.scores = scores


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill("solid", fgColor="DCFCE7")   # green
MISS_FILL = PatternFill("solid", fgColor="FEE2E2")    # red
EXTRA_FILL = PatternFill("solid", fgColor="FEF3C7")   # amber
NEUTRAL_FILL = PatternFill("solid", fgColor="F3F4F6") # light gray


def _autosize(ws, cols: list[int], max_width: int = 60) -> None:
    for c in cols:
        max_len = 10
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            v = row[0]
            if v is None:
                continue
            s = str(v)
            if "\n" in s:
                s = max(s.splitlines(), key=len)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, max_width)


def write_scorecard(out_dir: Path, reports: list[PilotReport]) -> Path:
    wb = Workbook()
    # First sheet: high-level summary
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Pilot", "Run ID", "Stage", "Truth comments",
                    "V4 total", "V4 Fail+NR", "Categories truth hit",
                    "Cats matched", "Cats missed", "Cats extra", "Recall %", "Notes"])
    for c in range(1, 13):
        cell = summary.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        truth_cats = sum(1 for s in rep.scores if s.has_truth)
        matched_cats = sum(1 for s in rep.scores if s.matched)
        missed_cats = sum(1 for s in rep.scores if s.missed)
        extra_cats = sum(1 for s in rep.scores if s.extra)
        recall = (matched_cats / truth_cats * 100) if truth_cats else 0.0
        v4_total = sum(s.v4_total for s in rep.scores)
        v4_nr = sum(s.v4_fail_or_nr for s in rep.scores)
        summary.append([
            rep.pilot.name,
            (rep.run_id_used or "")[:12],
            rep.pilot.stage,
            len(rep.truth),
            v4_total,
            v4_nr,
            truth_cats,
            matched_cats,
            missed_cats,
            extra_cats,
            round(recall, 1),
            "; ".join(rep.notes),
        ])
    _autosize(summary, list(range(1, 13)))
    summary.freeze_panes = "A2"

    # One sheet per pilot with per-category breakdown
    for rep in reports:
        safe = re.sub(r"[^A-Za-z0-9 ]+", "", rep.pilot.name)[:28]
        ws = wb.create_sheet(title=safe or rep.pilot.project_tag)
        ws.append(["Category", "Truth", "V4 total", "V4 Fail+NR",
                   "Match?", "Miss?", "Extra?", "Verdict"])
        for c in range(1, 9):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for s in rep.scores:
            verdict = []
            if s.matched:
                verdict.append("MATCH")
            if s.missed:
                verdict.append("MISS")
            if s.extra:
                verdict.append("EXTRA")
            if not verdict:
                verdict.append("-")
            ws.append([
                s.category, s.truth_count, s.v4_total, s.v4_fail_or_nr,
                "Y" if s.matched else "",
                "Y" if s.missed else "",
                "Y" if s.extra else "",
                ",".join(verdict),
            ])
            last_row = ws.max_row
            # color the verdict cell
            fill = NEUTRAL_FILL
            if s.matched:
                fill = MATCH_FILL
            elif s.missed:
                fill = MISS_FILL
            elif s.extra:
                fill = EXTRA_FILL
            for c in range(1, 9):
                ws.cell(row=last_row, column=c).fill = fill
        _autosize(ws, list(range(1, 9)))
        ws.freeze_panes = "A2"

    path = out_dir / "scorecard.xlsx"
    wb.save(path)
    return path


def write_ground_truth(out_dir: Path, reports: list[PilotReport]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "GroundTruth"
    ws.append(["Pilot", "xlsx row", "Comment ID", "Sheet (raw)",
               "V4 categories", "Severity", "Status", "Stage", "Text"])
    for c in range(1, 10):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        for t in rep.truth:
            ws.append([
                t.pilot, t.row, t.comment_id, t.sheet_raw,
                ", ".join(t.v4_categories) or "(unmapped)",
                t.severity, t.status, t.stage, t.text,
            ])
            row = ws.max_row
            if not t.v4_categories:
                for c in range(1, 10):
                    ws.cell(row=row, column=c).fill = EXTRA_FILL
    _autosize(ws, list(range(1, 10)), max_width=100)
    ws.column_dimensions["I"].width = 100
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    path = out_dir / "ground_truth.xlsx"
    wb.save(path)
    return path


def write_v4_emissions(out_dir: Path, reports: list[PilotReport]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "V4_Emissions"
    ws.append(["Pilot", "Run ID", "Category", "Rule Key", "Severity", "Status",
               "Page", "Title", "Description"])
    for c in range(1, 10):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rep in reports:
        for f in rep.v4:
            ws.append([
                f.pilot, (f.run_id or "")[:12], f.category, f.rule_key,
                f.severity, f.status, f.page_number, f.title, f.description,
            ])
    _autosize(ws, list(range(1, 10)), max_width=80)
    ws.column_dimensions["I"].width = 80
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    path = out_dir / "v4_emissions.xlsx"
    wb.save(path)
    return path


def write_diff_json(out_dir: Path, reports: list[PilotReport]) -> Path:
    serial = []
    for rep in reports:
        serial.append({
            "pilot": rep.pilot.name,
            "pdf": rep.pilot.pdf_name,
            "stage": rep.pilot.stage,
            "run_id": rep.run_id_used,
            "truth_count": len(rep.truth),
            "v4_total": sum(s.v4_total for s in rep.scores),
            "v4_fail_or_nr": sum(s.v4_fail_or_nr for s in rep.scores),
            "categories": [asdict(s) for s in rep.scores],
            "notes": rep.notes,
        })
    path = out_dir / "diff.json"
    path.write_text(json.dumps(serial, indent=2))
    return path


def write_summary_md(out_dir: Path, reports: list[PilotReport], mode: str) -> Path:
    lines = [
        "# V4 Regression — Summary",
        "",
        f"_Run timestamp_: `{datetime.now(timezone.utc).isoformat()}`",
        f"_Mode_: `{mode}`",
        "",
        "Ground truth comes from reviewer/customer comment logs in",
        "`2026-04-16-AI QC/Commented PDFs/`. For each pilot we compare the V4",
        "engine's emissions (from SQLite) against those labeled comments at",
        "V4-category granularity.",
        "",
        "## Headline",
        "",
        "| Pilot | Stage | Truth | V4 total / NR+Fail | Cats matched / missed / extra | Recall |",
        "|---|---|---|---|---|---|",
    ]
    for rep in reports:
        truth_cats = sum(1 for s in rep.scores if s.has_truth)
        matched_cats = sum(1 for s in rep.scores if s.matched)
        missed_cats = sum(1 for s in rep.scores if s.missed)
        extra_cats = sum(1 for s in rep.scores if s.extra)
        recall = (matched_cats / truth_cats * 100) if truth_cats else 0.0
        v4_total = sum(s.v4_total for s in rep.scores)
        v4_nr = sum(s.v4_fail_or_nr for s in rep.scores)
        lines.append(
            f"| {rep.pilot.name} | {rep.pilot.stage} | {len(rep.truth)} | "
            f"{v4_total} / {v4_nr} | {matched_cats}/{missed_cats}/{extra_cats} | "
            f"{recall:.0f}% |"
        )
    lines.append("")

    for rep in reports:
        lines.append(f"## {rep.pilot.name}")
        lines.append("")
        if rep.pilot.truth_caveat:
            lines.append(f"> **Truth-source caveat ({rep.pilot.truth_mode}):** "
                         f"{rep.pilot.truth_caveat}")
            lines.append("")
        if rep.notes:
            for n in rep.notes:
                lines.append(f"> {n}")
            lines.append("")
        lines.append(f"- xlsx: `{rep.pilot.xlsx}`")
        lines.append(f"- PDF : `{rep.pilot.pdf_name}`  (stage `{rep.pilot.stage}`)")
        lines.append(f"- truth mode: `{rep.pilot.truth_mode}`")
        if rep.run_id_used:
            lines.append(f"- V4 run: `{rep.run_id_used}`")
        else:
            lines.append("- V4 run: **NONE — no cached run and mode did not analyze**")
        lines.append("")

        # Category breakdown
        lines.append("| Category | Truth | V4 total | V4 NR+Fail | Verdict |")
        lines.append("|---|---|---|---|---|")
        for s in rep.scores:
            verdict_bits = []
            if s.matched:
                verdict_bits.append("MATCH")
            if s.missed:
                verdict_bits.append("MISS")
            if s.extra:
                verdict_bits.append("EXTRA")
            if not verdict_bits:
                continue  # no signal in either direction; skip
            lines.append(
                f"| {s.category} | {s.truth_count} | {s.v4_total} | "
                f"{s.v4_fail_or_nr} | {', '.join(verdict_bits)} |"
            )
        lines.append("")

    lines.append("## How to read this")
    lines.append("")
    lines.append(
        "- **MATCH** means at least one reviewer comment on that category "
        "AND at least one V4 Fail/Needs-Review finding on the same category.")
    lines.append(
        "- **MISS** means a reviewer flagged the category but V4 emitted "
        "no Fail/Needs-Review findings there. Investigate prompt or rule coverage.")
    lines.append(
        "- **EXTRA** means V4 emitted findings but the reviewer said nothing. "
        "Could be a real catch (reviewer missed it) or a false positive.")
    lines.append(
        "- Scoring is at CATEGORY granularity for now. Per-comment matching "
        "(does this V4 finding address *this specific* reviewer comment?) is "
        "the next iteration — requires text similarity.")

    path = out_dir / "summary.md"
    path.write_text("\n".join(lines))
    return path


# ---------------------------------------------------------------------------
# Run-V4 mode (optional)
# ---------------------------------------------------------------------------


def maybe_run_v4(pilot: Pilot) -> str | None:
    """Analyze the pilot PDF via the existing backend pipeline. Returns new run_id.

    Only invoked in --mode run-v4. Requires a valid Gemini/OpenAI API key and
    RULES_FILE=rules_v4_draft.yaml in the environment. Keeps imports deferred so
    the truth-only and use-cached modes don't need the full app wired up.
    """
    # Deferred imports keep the script usable without backend env.
    import os
    os.environ.setdefault("RULES_FILE", "rules_v4_draft.yaml")

    sys.path.insert(0, str(BACKEND))
    import fitz
    from app import analyzer  # type: ignore
    from app import gemini_analyzer  # type: ignore
    from app.db import init_db, insert_run  # type: ignore

    init_db()

    pdf_path = _find_pdf_on_disk(pilot)
    if not pdf_path:
        raise FileNotFoundError(f"PDF not found for pilot {pilot.name}: {pilot.pdf_name}")

    import uuid
    run_id = str(uuid.uuid4())
    run_dir = BACKEND / "data" / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    dst = run_dir / pdf_path.name
    dst.write_bytes(pdf_path.read_bytes())

    doc = fitz.open(dst)
    pages = analyzer.extract_pages(doc)

    # Minimal project_details shim; analyzer expects a dict.
    project_details = {
        "project_name": pilot.name,
        "design_stage": pilot.stage,
    }

    # Run the full analyze pipeline (same as POST /api/analyze).
    issues, category_summaries, status_counts, _summary = analyzer.analyze_pdf(
        doc=doc,
        pages=pages,
        run_id=run_id,
        run_dir=run_dir,
        project_details=project_details,
        supporting_docs=None,
    )
    doc.close()

    now = datetime.now(timezone.utc).isoformat()
    run_record = {
        "id": run_id,
        "project_name": pilot.name,
        "original_filename": pdf_path.name,
        "created_at": now,
        "pdf_path": str(dst),
        "page_count": len(pages),
        "summary": _summary,
        "status_counts": status_counts,
        "categories": category_summaries,
        "project_details": project_details,
    }
    insert_run(run_record, issues)
    return run_id


def _find_pdf_on_disk(pilot: Pilot) -> Path | None:
    """Locate the pilot's PDF. Checks cached run dirs and Commented PDFs/."""
    # Cached run dir?
    for run_dir in (BACKEND / "data" / "runs").glob("*"):
        cand = run_dir / pilot.pdf_name
        if cand.exists():
            return cand
    # Commented PDFs folder?
    cand = COMMENTED_DIR / pilot.pdf_name
    if cand.exists():
        return cand
    # Loose search
    for p in COMMENTED_DIR.glob("*.pdf"):
        if pilot.pdf_name.lower() in p.name.lower():
            return p
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def process_pilot(pilot: Pilot, mode: str) -> PilotReport:
    rep = PilotReport(pilot=pilot)
    # Ground truth is always loaded.
    try:
        rep.truth = load_ground_truth(pilot)
    except Exception as e:
        rep.notes.append(f"Ground-truth load FAILED: {e}")
        return rep

    unmapped = [t for t in rep.truth if not t.v4_categories]
    if unmapped:
        rep.notes.append(
            f"{len(unmapped)}/{len(rep.truth)} comments could not be mapped to "
            f"a V4 category (see ground_truth.xlsx, amber rows)."
        )

    if mode == "truth-only":
        rep.notes.append("Mode=truth-only: skipped V4 lookup.")
        score(rep)
        return rep

    # Try cached run.
    run_id = None
    if DB_PATH.exists():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            run_id = find_run_id(conn, pilot)
        finally:
            conn.close()

    if run_id is None and mode == "run-v4":
        rep.notes.append(f"No cached run; analyzing PDF via V4 (mode=run-v4)...")
        try:
            run_id = maybe_run_v4(pilot)
            rep.notes.append(f"Analyzed — new run_id={run_id[:12]}")
        except Exception as e:
            rep.notes.append(f"V4 analyze FAILED: {e}")
            score(rep)
            return rep

    if run_id is None:
        rep.notes.append(
            f"No cached run for '{pilot.pdf_name}' and mode={mode} won't analyze. "
            "Use --mode run-v4 or analyze via the UI first."
        )
        score(rep)
        return rep

    rep.run_id_used = run_id
    rep.v4 = load_v4_findings(pilot, run_id)
    if not rep.v4:
        rep.notes.append(f"Run {run_id[:12]} has zero issues in SQLite.")
    score(rep)
    return rep


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mode", choices=("truth-only", "use-cached", "run-v4"),
                    default="truth-only",
                    help="What to compare against V4.")
    ap.add_argument("--pilots", nargs="*", default=None,
                    help="Filter by pilot project_tag (e.g. --pilots bishop cottonwood).")
    ap.add_argument("--out-dir", default=None,
                    help="Override output directory.")
    args = ap.parse_args()

    pilots = PILOTS
    if args.pilots:
        keep = {p.lower() for p in args.pilots}
        pilots = [p for p in PILOTS if p.project_tag.lower() in keep]
        if not pilots:
            print(f"No pilots match {args.pilots!r}. Known tags: {[p.project_tag for p in PILOTS]}",
                  file=sys.stderr)
            return 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir) if args.out_dir else REGRESSION_DIR / ts
    out_dir.mkdir(parents=True, exist_ok=True)

    reports: list[PilotReport] = []
    for pilot in pilots:
        print(f"[*] {pilot.name} ({pilot.project_tag})")
        rep = process_pilot(pilot, args.mode)
        reports.append(rep)
        truth_n = len(rep.truth)
        v4_n = len(rep.v4)
        print(f"    truth={truth_n} v4={v4_n} notes={rep.notes}")

    scorecard = write_scorecard(out_dir, reports)
    gt = write_ground_truth(out_dir, reports)
    em = write_v4_emissions(out_dir, reports)
    dj = write_diff_json(out_dir, reports)
    sm = write_summary_md(out_dir, reports, mode=args.mode)

    print()
    print(f"[OK] Output directory: {out_dir}")
    print(f"     scorecard     : {scorecard.name}")
    print(f"     ground truth  : {gt.name}")
    print(f"     v4 emissions  : {em.name}")
    print(f"     diff json     : {dj.name}")
    print(f"     summary.md    : {sm.name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
help="What to compare against V4.")
    ap.add_argument("--pilots", nargs="*", default=None,
                    help="Filter by pilot project_tag (e.g. --pilots bishop cottonwood).")
    ap.add_argument("--out-dir", default=None,
                    help="Override output directory.")
    args = ap.parse_args()

    pilots = PILOTS
    if args.pilots:
        keep = {p.lower() for p in args.pilots}
        pilots = [p for p in PILOTS if p.project_tag.lower() in keep]
        if not pilots:
            print(f"No pilots match {args.pilots\!r}. Known tags: {[p.project_tag for p in PILOTS]}",
                  file=sys.stderr)
            return 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir) if args.out_dir else REGRESSION_DIR / ts
    out_dir.mkdir(parents=True, exist_ok=True)

    reports: list[PilotReport] = []
    for pilot in pilots:
        print(f"[*] {pilot.name} ({pilot.project_tag})")
        rep = process_pilot(pilot, args.mode)
        reports.append(rep)
        truth_n = len(rep.truth)
        v4_n = len(rep.v4)
        print(f"    truth={truth_n} v4={v4_n} notes={rep.notes}")

    scorecard = write_scorecard(out_dir, reports)
    gt = write_ground_truth(out_dir, reports)
    em = write_v4_emissions(out_dir, reports)
    dj = write_diff_json(out_dir, reports)
    sm = write_summary_md(out_dir, reports, mode=args.mode)

    print()
    print(f"[OK] Output directory: {out_dir}")
    print(f"     scorecard     : {scorecard.name}")
    print(f"     ground truth  : {gt.name}")
    print(f"     v4 emissions  : {em.name}")
    print(f"     diff json     : {dj.name}")
    print(f"     summary.md    : {sm.name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
 tags: {[p.project_tag for p in PILOTS]}",
                  file=sys.stderr)
            return 2

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir) if args.out_dir else REGRESSION_DIR / ts
    out_dir.mkdir(parents=True, exist_ok=True)

    reports: list[PilotReport] = []
    for pilot in pilots:
        print(f"[*] {pilot.name} ({pilot.project_tag})")
        rep = process_pilot(pilot, args.mode)
        reports.append(rep)
        truth_n = len(rep.truth)
        v4_n = len(rep.v4)
        print(f"    truth={truth_n} v4={v4_n} notes={rep.notes}")

    scorecard = write_scorecard(out_dir, reports)
    gt = write_ground_truth(out_dir, reports)
    em = write_v4_emissions(out_dir, reports)
    dj = write_diff_json(out_dir, reports)
    sm = write_summary_md(out_dir, reports, mode=args.mode)

    print()
    print(f"[OK] Output directory: {out_dir}")
    print(f"     scorecard     : {scorecard.name}")
    print(f"     ground truth  : {gt.name}")
    print(f"     v4 emissions  : {em.name}")
    print(f"     diff json     : {dj.name}")
    print(f"     summary.md    : {sm.name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
    print(f"     diff json     : {dj.name}")
    print(f"     summary.md    : {sm.name}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
