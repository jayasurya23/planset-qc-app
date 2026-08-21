from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CASTILLO_RED = "AD1F2B"
CASTILLO_DARK = "333132"
CASTILLO_GREEN = "278747"
CASTILLO_GREY = "E6E7E8"
CASTILLO_LIGHT_RED = "F8E0E2"
CASTILLO_BLUE = "E8F0FE"


# ---------------------------------------------------------------------------
# Due Diligence Template
# ---------------------------------------------------------------------------

_DD_SECTIONS: list[tuple[str, list[tuple[str, str, str]]]] = [
    # (Section Name, [(field_key, Label, Hint/Example), ...])
    ("PROJECT INFORMATION", [
        ("project_name", "Project Name", "e.g. Sunfield Solar Farm"),
        ("project_address", "Project Address", "Full street address"),
        ("site_coordinates", "Site Coordinates", "(latitude, longitude)  e.g. (39.7817, -89.6501)"),
        ("county", "County", ""),
        ("state", "State", "2-letter code  e.g. IL"),
        ("parcel_id", "Parcel ID", "County assessor parcel number"),
        ("building_codes", "Applicable Building Codes", "NEC 2020, IBC 2021, state/county amendments"),
        ("der_number", "DER Number (if Ameren)", "Distributed Energy Resource ID"),
    ]),
    ("OWNER", [
        ("owner_name", "Owner / Client Name", ""),
        ("owner_address", "Owner Address", ""),
        ("owner_phone", "Owner Phone", ""),
    ]),
    ("EPC CONTRACTOR", [
        ("epc_name", "EPC Company Name", ""),
        ("epc_address", "EPC Address", ""),
        ("epc_phone", "EPC Phone", ""),
    ]),
    ("ENGINEERING", [
        ("eor_name", "Engineer of Record (EOR)", "Full name"),
        ("eor_license", "EOR License Number", "PE license #"),
        ("eor_state", "EOR License State", "State where PE is licensed"),
        ("checker_name", "Checker Name", "QC checker assigned"),
        ("designer_name", "Designer Name", "Electrical designer"),
    ]),
    ("PV MODULE", [
        ("module_make", "Module Manufacturer", "e.g. Canadian Solar, LONGi, Trina"),
        ("module_model", "Module Model Number", "e.g. CS6W-550MS"),
        ("module_stc_watts", "STC Rating (W)", "e.g. 550"),
        ("module_voc", "Open Circuit Voltage Voc (V)", "From module datasheet"),
        ("module_vmp", "Max Power Voltage Vmp (V)", "From module datasheet"),
        ("module_isc", "Short Circuit Current Isc (A)", "From module datasheet"),
        ("module_imp", "Max Power Current Imp (A)", "From module datasheet"),
        ("module_temp_coeff_voc", "Temp Coefficient of Voc (%/\u00b0C)", "e.g. -0.27"),
        ("module_temp_coeff_isc", "Temp Coefficient of Isc (%/\u00b0C)", "e.g. 0.05"),
        ("is_bifacial", "Bifacial Module?", "Yes / No"),
    ]),
    ("INVERTER", [
        ("inverter_make", "Inverter Manufacturer", "e.g. SMA, Sungrow, Power Electronics"),
        ("inverter_model", "Inverter Model Number", ""),
        ("inverter_kva", "Inverter kVA Rating", "From inverter datasheet"),
        ("inverter_kw", "Inverter kW Rating", "From inverter datasheet"),
        ("inverter_max_vdc", "Max DC Input Voltage (V)", "From inverter datasheet"),
        ("inverter_mppt_range", "MPPT Voltage Range (V)", "e.g. 600\u20131500"),
        ("inverter_quantity", "Inverter Quantity", "Total count for the project"),
    ]),
    ("SYSTEM CONFIGURATION", [
        ("string_size", "String Size (modules/string)", "Modules per string"),
        ("string_quantity", "Total String Quantity", "Across entire project"),
        ("total_dc_kw", "Total DC System Size (kW)", ""),
        ("total_ac_kva", "Total AC System Size (kVA)", ""),
        ("dc_ac_ratio", "DC/AC Ratio", "e.g. 1.30"),
    ]),
    ("RACKING / MOUNTING", [
        ("racking_make", "Racking Manufacturer", "e.g. NEXTracker, GameChange"),
        ("racking_model", "Racking Model", ""),
        ("racking_type", "Racking Type", "Fixed Tilt / Single-Axis Tracker"),
        ("pitch", "Pitch (ft)", "Row-to-row pitch"),
        ("interrow_spacing", "Interrow Spacing (ft)", "Clear spacing between rows"),
        ("gcr", "Ground Coverage Ratio (GCR)", "e.g. 0.35"),
        ("tilt_angle", "Tilt Angle / Tracker Rotation (\u00b0)", ""),
        ("azimuth", "Azimuth (\u00b0)", "180 = due south"),
    ]),
    ("TRANSFORMER", [
        ("transformer_kva", "Transformer kVA Rating", "Must be \u2265 total inverter kVA"),
        ("transformer_primary_voltage", "Primary Voltage (V)", "MV side  e.g. 12470, 34500"),
        ("transformer_secondary_voltage", "Secondary Voltage (V)", "LV side  e.g. 480, 600"),
        ("transformer_winding_config", "Winding Configuration", "e.g. Delta\u2013Wye Grounded"),
        ("transformer_impedance", "Impedance Z%", "e.g. 5.75"),
        ("transformer_bil", "BIL (kV)", "e.g. 95 for 15 kV class"),
    ]),
    ("UTILITY / POINT OF INTERCONNECTION", [
        ("utility_name", "Utility Name", ""),
        ("utility_feeder", "Feeder Name / Number", ""),
        ("is_ngrid", "Is Utility NGrid?", "Yes / No \u2014 NGrid requires specific relay tables"),
        ("poi_voltage", "POI Voltage (V)", "e.g. 12470Y/7200"),
        ("feeder_grounding", "Feeder Grounding", "e.g. 4-wire multigrounded neutral"),
        ("fault_current", "Available Fault Current (kA)", "From utility study"),
        ("ieee_category", "IEEE 1547 Category", "I / II / III \u2014 default I if not specified in CESIR"),
    ]),
    ("MV EQUIPMENT / PROTECTION", [
        ("recloser_make", "Recloser Make", "e.g. Tavrida"),
        ("recloser_continuous_a", "Recloser Continuous Rating (A)", "e.g. 630"),
        ("recloser_interrupting_ka", "Recloser Interrupting Rating (kA)", "e.g. 12.5"),
        ("ct_ratio", "Metering CT Ratio", "Default 300:1 if no submittal"),
        ("vt_ratio", "Metering VT Ratio", "Default 234.5 for Tavrida recloser"),
        ("meter_accuracy_class", "Meter Accuracy Class", "0.3 for revenue metering"),
        ("surge_arrestor_mcov", "Surge Arrestor MCOV (kV)", "e.g. 8.4 for 12470Y/7200"),
    ]),
    ("DESIGN TEMPERATURES", [
        ("design_temp_low_c", "Design Low Temperature (\u00b0C)", "For Voc cold calc  e.g. -20"),
        ("design_temp_high_c", "Design High Temperature (\u00b0C)", "For Vmp hot calc  e.g. 45"),
        ("ambient_temp_c", "Ambient Temperature (\u00b0C)", "For cable derating  e.g. 35"),
    ]),
    ("ADDITIONAL NOTES & ATTACHMENTS", [
        ("special_notes", "Special Notes / Requirements", "Any deviations, utility-specific requirements, or design constraints"),
    ]),
]


def build_due_diligence_template() -> Workbook:
    """Create a formatted Excel due-diligence template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Due Diligence"

    hdr_fill = PatternFill("solid", fgColor=CASTILLO_RED)
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    section_fill = PatternFill("solid", fgColor=CASTILLO_DARK)
    section_font = Font(color="FFFFFF", bold=True, size=11)
    hint_font = Font(color="888888", italic=True, size=10)
    value_fill = PatternFill("solid", fgColor=CASTILLO_BLUE)
    label_font = Font(bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="center")

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Castillo Engineering \u2014 Project Due Diligence Sheet"
    ws["A1"].font = Font(size=16, bold=True, color=CASTILLO_RED)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Fill in all known values before submitting the planset for QC review. "
        "Attach equipment datasheets, interconnection agreements, CESIR, and impact studies."
    )
    ws["A2"].font = Font(size=10, color="666666")
    ws["A2"].alignment = wrap
    ws.row_dimensions[2].height = 32

    # Column headers
    row = 4
    for col, (hdr, w) in enumerate([("Section / Field", 40), ("Value", 35), ("Example / Hint", 40), ("Source Document", 22)], 1):
        c = ws.cell(row=row, column=col, value=hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = "A5"
    row += 1

    for section_name, fields in _DD_SECTIONS:
        # Section header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=section_name)
        c.fill = section_fill
        c.font = section_font
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1

        for _key, label, hint in fields:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).alignment = wrap

            # Value cell — highlighted so it's clear what to fill
            vc = ws.cell(row=row, column=2, value="")
            vc.fill = value_fill
            vc.alignment = wrap

            # Hint
            hc = ws.cell(row=row, column=3, value=hint)
            hc.font = hint_font
            hc.alignment = wrap

            # Source column
            ws.cell(row=row, column=4, value="").alignment = wrap

            ws.row_dimensions[row].height = 22
            row += 1

    # Footer
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=(
        "Prepared by: _________________   Date: _________________   "
        "Reviewed by: _________________   Date: _________________"
    )).font = Font(size=10, color="666666")
    ws.row_dimensions[row].height = 28

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=(
        "Castillo Engineering Services LLC  |  1060 Maitland Center Commons Blvd, Suite 270, Maitland, FL 32751  |  (407) 289-2575"
    )).font = Font(size=9, color="999999")

    ws.print_title_rows = "4:4"
    return wb


def export_due_diligence(output_path: Path) -> Path:
    wb = build_due_diligence_template()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


STATUS_FILL = {
    "Pass": "E6F4EA",
    "Fail": "FCE8E6",
    "Needs Review": "FFF4CC",
    "Deferred": "EFF8FF",
    "Overridden / Accepted by QC Engineer": "E8F0FE",
}


def _format_locations(issue: dict) -> str | None:
    """For a multi-location (cross-sheet conflict) finding, render the
    conflicting sheet/page + value pairs as a single cell string, e.g.::

        E-100 p7: 2250 kVA; cover p1: 2000 kVA

    Returns None for normal single-location findings so the cell stays blank.
    """
    locations = issue.get("locations")
    if not locations or not isinstance(locations, list) or len(locations) < 2:
        return None
    parts: list[str] = []
    for loc in locations:
        if not isinstance(loc, dict):
            continue
        label = loc.get("source_label") or loc.get("sheet_number") or "?"
        page = loc.get("page_number")
        value = loc.get("value")
        page_str = f" p{page}" if page else ""
        parts.append(f"{label}{page_str}: {value}")
    return "; ".join(parts) if parts else None


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


SNIPPET_WIDTH_PX = 180  # embedded snippet width; column L is sized to suit


def build_workbook(run: dict) -> Workbook:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill("solid", fgColor=CASTILLO_RED)
    header_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill("solid", fgColor=CASTILLO_GREY)

    ws_summary["A1"] = "Castillo Planset QC Summary"
    ws_summary["A1"].font = Font(size=16, bold=True, color=CASTILLO_DARK)
    ws_summary["A3"] = "Project"
    ws_summary["B3"] = run["project_name"]
    ws_summary["A4"] = "Original file"
    ws_summary["B4"] = run["original_filename"]
    ws_summary["A5"] = "Created"
    ws_summary["B5"] = run["created_at"]

    ws_summary["A7"] = "Metric"
    ws_summary["B7"] = "Value"
    for cell in (ws_summary["A7"], ws_summary["B7"]):
        cell.fill = header_fill
        cell.font = header_font

    metrics = [
        ("PDF pages", run["summary"]["pdf_page_count"]),
        ("Indexed sheets", run["summary"]["indexed_sheet_count"]),
        ("Actual sheets", run["summary"]["actual_sheet_count"]),
        ("Pass", run["status_counts"].get("Pass", 0)),
        ("Fail", run["status_counts"].get("Fail", 0)),
        ("Needs Review", run["status_counts"].get("Needs Review", 0)),
        ("Deferred", run["status_counts"].get("Deferred", 0)),
        (
            "Overridden / Accepted by QC Engineer",
            run["status_counts"].get("Overridden / Accepted by QC Engineer", 0),
        ),
    ]
    for idx, (metric, value) in enumerate(metrics, start=8):
        ws_summary[f"A{idx}"] = metric
        ws_summary[f"B{idx}"] = value

    ws_summary["D3"] = "Category"
    ws_summary["E3"] = "Total"
    ws_summary["F3"] = "Pass"
    ws_summary["G3"] = "Fail"
    ws_summary["H3"] = "Needs Review"
    ws_summary["I3"] = "Deferred"
    ws_summary["J3"] = "Overridden"
    for cell in ws_summary["D3:J3"][0]:
        cell.fill = header_fill
        cell.font = header_font

    for idx, category in enumerate(run["categories"], start=4):
        ws_summary[f"D{idx}"] = category["name"]
        ws_summary[f"E{idx}"] = category["total"]
        ws_summary[f"F{idx}"] = category.get("Pass", 0)
        ws_summary[f"G{idx}"] = category.get("Fail", 0)
        ws_summary[f"H{idx}"] = category.get("Needs Review", 0)
        ws_summary[f"I{idx}"] = category.get("Deferred", 0)
        ws_summary[f"J{idx}"] = category.get("Overridden / Accepted by QC Engineer", 0)

    autosize_columns(ws_summary)

    ws_issues = wb.create_sheet("Issues")
    headers = [
        "Category",
        "Checklist Item",
        "Status",
        "Auto Status",
        "Severity",
        "Page",
        "Confidence",
        "Evidence",
        "Description",
        "Override Comment",
        "Locations",
        "Snippet",
    ]
    ws_issues.append(headers)
    for cell in ws_issues[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws_issues.freeze_panes = "A2"

    row_idx = 2
    for issue in run["issues"]:
        ws_issues.cell(row=row_idx, column=1, value=issue["category"])
        ws_issues.cell(row=row_idx, column=2, value=issue["title"])
        ws_issues.cell(row=row_idx, column=3, value=issue["status"])
        ws_issues.cell(row=row_idx, column=4, value=issue["auto_status"])
        ws_issues.cell(row=row_idx, column=5, value=issue["severity"])
        ws_issues.cell(row=row_idx, column=6, value=issue.get("page_number"))
        ws_issues.cell(row=row_idx, column=7, value=issue["confidence"])
        ws_issues.cell(row=row_idx, column=8, value=issue.get("evidence"))
        ws_issues.cell(row=row_idx, column=9, value=issue.get("description"))
        ws_issues.cell(row=row_idx, column=10, value=issue.get("override_comment"))
        # Locations — cross-sheet conflict findings list every conflicting
        # sheet/page + value pair here; blank for normal single-location ones.
        ws_issues.cell(row=row_idx, column=11, value=_format_locations(issue))

        fill_color = STATUS_FILL.get(issue["status"], "FFFFFF")
        ws_issues.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=fill_color)

        for col in range(1, 12):
            ws_issues.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        snippet_path = issue.get("snippet_path")
        if snippet_path and Path(snippet_path).exists():
            img = XLImage(snippet_path)
            # Capture the natural size BEFORE resizing. Reading img.width on
            # the next line after assigning it makes the ratio 1.0, so the
            # height stayed at the source pixel height and every snippet came
            # out stretched -- 4.5x on a focused crop, 28.8x on a full page
            # (5184x3456 embedded at 180x3456 instead of 180x120), burying the
            # findings table the image is there to illustrate.
            natural_w, natural_h = img.width, img.height
            img.width = SNIPPET_WIDTH_PX
            img.height = (int(natural_h * (SNIPPET_WIDTH_PX / natural_w))
                          if natural_w else 120)
            anchor = f"L{row_idx}"
            ws_issues.add_image(img, anchor)
            # Excel row heights are points, image heights pixels: 1 px = 0.75 pt.
            needed = img.height * 0.75 + 6
            ws_issues.row_dimensions[row_idx].height = max(
                ws_issues.row_dimensions[row_idx].height or 15, 110, needed)
        row_idx += 1

    for col in range(1, 13):
        ws_issues.column_dimensions[get_column_letter(col)].width = 24
    ws_issues.column_dimensions["H"].width = 40
    ws_issues.column_dimensions["I"].width = 40
    ws_issues.column_dimensions["K"].width = 48
    return wb


def export_run_to_excel(run: dict, output_path: Path) -> Path:
    wb = build_workbook(run)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
