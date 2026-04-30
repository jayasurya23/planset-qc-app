"""Auto-detect column meanings in reviewer-comment xlsx logs.

Used by the V4 regression harness (scripts/regression_all_plansets.py) to scale
ground-truth ingestion from 3 hand-configured pilots to all 84 matched logs
without hand-writing a Pilot() entry per template.

The sniffer searches the first ~15 rows for the header row by scoring each
candidate row against known patterns for:

    comment column (high weight)       e.g. "PPE Engineer Comment 1"
    sheet/page column                  e.g. "Sheet Page / Location on Document"
    severity                           e.g. "Severity"
    status                             e.g. "Status"

Returns a SniffResult or raises if no plausible header is found.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Priority-ordered patterns (substring match, case-insensitive)
# ---------------------------------------------------------------------------

COMMENT_COL_PATTERNS: list[str] = [
    # PPE / PVP Engineer Comment family (most common)
    "ppe engineer comment 1",
    "ppe engineer comment #1",
    "pvp engineer comment 1",
    "pvp engineer comment #1",
    # Owner-dialog templates (Grass River, Brookhaven)
    "owner comment",
    "pseg-li comments",
    "county comment 1",
    # DNV-style "Issues list" and misc
    "issues list",
    "question/comment",
    "comment / question",
    "madison notes",
    "drawing comment",
    "reviewer comment",
    "eor comment",
    "description",
    "issue",
    # PPE Engineer Comment N (fallback to any numbered)
    "ppe engineer comment",
    "pvp engineer comment",
    "engineer comment",
    "ppe comment #1",
    "ppe comment 1",
    "ppe comment",
    # Generic fallback — least specific
    "contractor comment",
    "comment",
]

SHEET_COL_PATTERNS: list[str] = [
    "sheet page / location on document",
    "sheet page",
    "sheet ref",
    "sheet number",
    "drawing sheet",
    "sheet #",
    "sheet title",
    "sheet desription",  # (sic) typo seen in source files
    "sheet",
    "location on document",
    "section",
    "area of system",
    "sn or location",
    "drawing for review",
    "page number",   # Bishop template
    "page",
    "item",          # Brookhaven
]

SEVERITY_PATTERNS = ["severity", "priority", "rating"]
STATUS_PATTERNS = ["status"]


# Exclusion patterns to avoid picking things like "Comment ID", "Comment Date",
# "Coast Response/Action", "Date Closed"  as the comment column.
COMMENT_EXCLUDE = (
    "comment id", "comment number", "comment date", "comment by",
    "response", "date", "reviewer name", "reviewer company",
    "commenter",
)


@dataclass
class SniffResult:
    sheet_name: str
    header_row: int               # 1-indexed
    headers: list[str]
    comment_col: Optional[str]
    sheet_col: Optional[str]
    severity_col: Optional[str]
    status_col: Optional[str]
    score: int                    # confidence: higher = better
    warnings: list[str]

    @property
    def viable(self) -> bool:
        """Is there enough signal to treat this xlsx as structured ground truth?"""
        return self.comment_col is not None and self.sheet_col is not None


def _find_col(headers: list[str], patterns: list[str],
              exclude: tuple[str, ...] = ()) -> Optional[str]:
    lowered = [h.lower() for h in headers]
    for pat in patterns:
        for i, h in enumerate(lowered):
            if pat in h and not any(x in h for x in exclude):
                return headers[i]
    return None


def _score_header_row(headers: list[str]) -> int:
    """Heuristic score: prefer rows that look like column headers for a
    reviewer comment log."""
    lowered = [h.lower() for h in headers]
    score = 0
    for pat_list, weight in (
        (COMMENT_COL_PATTERNS, 3),
        (SHEET_COL_PATTERNS, 2),
        (SEVERITY_PATTERNS, 1),
        (STATUS_PATTERNS, 1),
    ):
        if any(pat in h for pat in pat_list for h in lowered):
            score += weight
    return score


def sniff_worksheet(
    ws,
    max_scan_rows: int = 15,
    max_cols: int = 40,
) -> SniffResult:
    """Inspect the first `max_scan_rows` of a worksheet and return the best
    candidate header row + its extracted column mappings.

    If no row scores > 0, returns a SniffResult with `viable=False` so the
    caller can skip the pilot gracefully.
    """
    rows = list(ws.iter_rows(values_only=True, max_row=max_scan_rows, max_col=max_cols))

    best_score = -1
    best_row_idx = 0
    best_headers: list[str] = []

    for i, row in enumerate(rows, start=1):
        headers = [str(v).strip() if v is not None else "" for v in row]
        non_empty = sum(1 for h in headers if h)
        if non_empty < 3:
            continue
        s = _score_header_row(headers)
        if s > best_score:
            best_score = s
            best_row_idx = i
            best_headers = headers

    warnings: list[str] = []
    if best_score <= 0:
        warnings.append(f"No plausible header row in first {max_scan_rows} rows.")

    comment = _find_col(best_headers, COMMENT_COL_PATTERNS, exclude=COMMENT_EXCLUDE)
    sheet = _find_col(best_headers, SHEET_COL_PATTERNS)
    severity = _find_col(best_headers, SEVERITY_PATTERNS)
    status = _find_col(best_headers, STATUS_PATTERNS)

    if comment is None:
        warnings.append("No comment column detected.")
    if sheet is None:
        warnings.append("No sheet/page column detected.")

    return SniffResult(
        sheet_name=ws.title,
        header_row=best_row_idx,
        headers=best_headers,
        comment_col=comment,
        sheet_col=sheet,
        severity_col=severity,
        status_col=status,
        score=max(best_score, 0),
        warnings=warnings,
    )


def sniff_xlsx(path: Path | str) -> SniffResult:
    """Open the xlsx (read-only) and sniff its *first* worksheet."""
    p = Path(path)
    wb = load_workbook(p, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return sniff_worksheet(ws)
    finally:
        wb.close()


__all__ = [
    "SniffResult",
    "sniff_worksheet",
    "sniff_xlsx",
    "COMMENT_COL_PATTERNS",
    "SHEET_COL_PATTERNS",
    "SEVERITY_PATTERNS",
    "STATUS_PATTERNS",
]
